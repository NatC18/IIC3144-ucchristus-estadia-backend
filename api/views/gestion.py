"""
Views para el modelo Gestion
"""

from datetime import datetime

from django.db.models import Count, Q
from django.http import HttpResponse
from django_filters import CharFilter, FilterSet
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from api.models import Gestion
from api.serializers import (
    GestionCreateSerializer,
    GestionListSerializer,
    GestionSerializer,
    GestionUpdateSerializer,
)


class GestionFilterSet(FilterSet):
    """
    Custom FilterSet para manejar filtros especiales como 'not_assigned'
    """

    usuario = CharFilter(method="filter_usuario")

    def filter_usuario(self, queryset, name, value):
        """
        Filtro personalizado para usuario
        Si value es 'not_assigned', retorna gestiones sin usuario asignado
        Si value es un UUID, retorna gestiones con ese usuario
        """
        if value == "not_assigned":
            return queryset.filter(usuario__isnull=True)
        elif value and value != "all":
            return queryset.filter(usuario__id=value)
        return queryset

    class Meta:
        model = Gestion
        fields = ["estado_gestion", "tipo_gestion", "episodio", "usuario"]


class GestionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión completa de gestiones

    Operaciones disponibles:
    - GET /api/gestiones/ - Listar gestiones
    - POST /api/gestiones/ - Crear gestión
    - GET /api/gestiones/{id}/ - Obtener gestión específica
    - PUT /api/gestiones/{id}/ - Actualizar gestión completa
    - PATCH /api/gestiones/{id}/ - Actualizar gestión parcial
    - DELETE /api/gestiones/{id}/ - Eliminar gestión
    """

    queryset = Gestion.objects.select_related(
        "episodio", "episodio__paciente", "usuario"
    ).all()
    serializer_class = GestionSerializer
    permission_classes = [IsAuthenticated]

    # Filtros y búsqueda
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = GestionFilterSet
    search_fields = ["tipo_gestion", "informe"]
    ordering_fields = ["fecha_inicio", "fecha_fin", "created_at"]
    ordering = ["-fecha_inicio"]

    def get_serializer_class(self):
        """
        Retorna el serializer apropiado según la acción
        """
        if self.action == "create":
            return GestionCreateSerializer
        elif self.action in ["update", "partial_update"]:
            return GestionUpdateSerializer
        elif self.action == "list":
            return GestionListSerializer
        return GestionSerializer

    @action(detail=False, methods=["get"])
    def pendientes(self, request):
        """
        Listar gestiones pendientes (INICIADA o EN_PROGRESO)
        GET /api/gestiones/pendientes/
        """
        gestiones_pendientes = self.get_queryset().filter(
            estado_gestion__in=["INICIADA", "EN_PROGRESO"]
        )
        serializer = self.get_serializer(gestiones_pendientes, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def estadisticas(self, request):
        """
        Estadísticas de gestiones para el dashboard
        GET /api/gestiones/estadisticas/
        """
        queryset = self.get_queryset()

        # Contar por estado
        por_estado = (
            queryset.values("estado_gestion")
            .annotate(cantidad=Count("id"))
            .order_by("estado_gestion")
        )

        # Contar por tipo de gestión
        por_tipo = (
            queryset.values("tipo_gestion")
            .annotate(cantidad=Count("id"))
            .order_by("-cantidad")
        )

        # Transformar tipo_gestion a nombres legibles
        tipo_gestion_data = []
        for item in por_tipo:
            # Obtener el display name usando el modelo
            tipo_code = item["tipo_gestion"]
            # Buscar el label del choice
            tipo_label = dict(Gestion.TIPO_GESTION_CHOICES).get(tipo_code, tipo_code)
            tipo_gestion_data.append(
                {"tipo_gestion": tipo_label, "cantidad": item["cantidad"]}
            )

        return Response(
            {
                "total_gestiones": queryset.count(),
                "por_estado": list(por_estado),
                "por_tipo": list(por_tipo),
                "por_tipo_gestion": tipo_gestion_data,
            }
        )

    @action(detail=False, methods=["get"])
    def tareas_pendientes(self, request):
        """
        Lista de tareas pendientes formateadas para el dashboard
        GET /api/gestiones/tareas_pendientes/
        """
        gestiones = (
            self.get_queryset()
            .filter(estado_gestion__in=["INICIADA", "EN_PROGRESO"])
            .select_related("episodio", "episodio__paciente")[:10]
        )

        # Mapeo de estados
        estado_map = {
            "INICIADA": "Abierta",
            "EN_PROGRESO": "En proceso",
            "COMPLETADA": "Cerrada",
            "CANCELADA": "Cancelada",
        }

        data = []
        for g in gestiones:
            data.append(
                {
                    "id": str(g.id),
                    "episodio": str(g.episodio.episodio_cmbd),
                    "tipo_gestion": g.get_tipo_gestion_display(),
                    "descripcion": g.informe
                    or f"Gestión de {g.get_tipo_gestion_display()}",
                    "estado": estado_map.get(g.estado_gestion, "Abierta"),
                    "fecha_inicio": (
                        g.fecha_inicio.isoformat() if g.fecha_inicio else None
                    ),
                }
            )

        return Response(data)

    @action(detail=False, methods=["get"], url_path="exportar-excel")
    def exportar_excel(self, request):
        """
        Exporta todas las gestiones a un archivo Excel
        GET /api/gestiones/exportar-excel/
        Requiere: openpyxl
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {
                    "error": "La librería openpyxl no está instalada. Use exportar-csv en su lugar."
                },
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Gestiones"

        # Estilo para encabezados
        header_fill = PatternFill(
            start_color="671E75", end_color="671E75", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # Encabezados
        headers = [
            "Episodio CMBD",
            "RUT Paciente",
            "Nombre Paciente",
            "Usuario Responsable",
            "Tipo de Gestión",
            "Estado Gestión",
            "Fecha Inicio",
            "Fecha Fin",
            "Duración (días)",
            "Informe",
            "Estado Traslado",
            "Tipo Traslado",
            "Motivo Traslado",
            "Centro Destinatario",
            "Tipo Solicitud Traslado",
            "Nivel Atención Traslado",
            "Motivo Rechazo",
            "Motivo Cancelación",
            "Fecha Finalización Traslado",
            "Fecha Creación",
            "Última Actualización",
        ]

        # Escribir encabezados con estilo
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Escribir datos
        gestiones = self.get_queryset()

        # Función helper para convertir datetime con timezone a naive
        def to_naive_datetime(dt):
            """Convierte datetime aware a naive para compatibilidad con Excel"""
            if dt is None:
                return None
            if hasattr(dt, "tzinfo") and dt.tzinfo is not None:
                # Convertir a la zona horaria local y quitar el tzinfo
                from django.utils import timezone

                local_dt = timezone.localtime(dt)
                return local_dt.replace(tzinfo=None)
            return dt

        for row_num, gestion in enumerate(gestiones, 2):
            data = [
                gestion.episodio.episodio_cmbd if gestion.episodio else "",
                (
                    gestion.episodio.paciente.rut
                    if gestion.episodio and gestion.episodio.paciente
                    else ""
                ),
                (
                    gestion.episodio.paciente.nombre
                    if gestion.episodio and gestion.episodio.paciente
                    else ""
                ),
                gestion.usuario.email if gestion.usuario else "",
                gestion.get_tipo_gestion_display(),
                gestion.get_estado_gestion_display(),
                to_naive_datetime(gestion.fecha_inicio),
                to_naive_datetime(gestion.fecha_fin),
                gestion.duracion_dias,
                gestion.informe or "",
                (
                    gestion.get_estado_traslado_display()
                    if gestion.estado_traslado
                    else ""
                ),
                (gestion.get_tipo_traslado_display() if gestion.tipo_traslado else ""),
                gestion.motivo_traslado or "",
                gestion.centro_destinatario or "",
                (
                    gestion.get_tipo_solicitud_traslado_display()
                    if gestion.tipo_solicitud_traslado
                    else ""
                ),
                (
                    gestion.get_nivel_atencion_traslado_display()
                    if gestion.nivel_atencion_traslado
                    else ""
                ),
                gestion.motivo_rechazo_traslado or "",
                gestion.motivo_cancelacion_traslado or "",
                to_naive_datetime(gestion.fecha_finalizacion_traslado),
                to_naive_datetime(gestion.created_at),
                to_naive_datetime(gestion.updated_at),
            ]

            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajustar ancho de columnas
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        # Crear respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = (
            f'attachment; filename="gestiones_{timestamp}.xlsx"'
        )

        wb.save(response)
        return response
